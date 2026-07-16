#!/usr/bin/env python3
"""
Snapchat Ad Audit Report Generator — Allformance
Usage: Called by Claude agent with parsed data. Not a standalone CLI.

This script provides building blocks for generating audit Excel reports.
The agent calls individual functions with real data extracted from
whatever format the user provides (CSV, text, screenshots, etc.)
"""

import os
from datetime import datetime
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

# ─── Brand Colors (Allformance) ───
COLORS = {
    "dark": "1A1A2E",
    "primary": "2163E7",
    "accent": "F59E0B",
    "green": "10B981",
    "red": "EF4444",
    "light_bg": "F8F9FC",
    "white": "FFFFFF",
    "gray": "6B7280",
    "light_gray": "E5E7EB",
    "purple": "9333EA",
}

# ─── Reusable Styles ───
def _font(bold=False, size=10, color="dark"):
    return Font(name="Calibri", bold=bold, size=size, color=COLORS.get(color, color))

def _fill(color="dark"):
    c = COLORS.get(color, color)
    return PatternFill(start_color=c, end_color=c, fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color=COLORS["light_gray"]),
    right=Side(style="thin", color=COLORS["light_gray"]),
    top=Side(style="thin", color=COLORS["light_gray"]),
    bottom=Side(style="thin", color=COLORS["light_gray"]),
)


class AuditReport:
    """Builds an audit Excel workbook sheet by sheet."""

    def __init__(self, client_name: str, period: str, output_dir: str):
        self.client_name = client_name
        self.period = period
        self.output_dir = output_dir
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)  # remove default sheet
        self._insights: dict[str, list[str]] = {}  # sheet_name -> insights

    # ─── Utility Methods ───

    def _style_header(self, ws, row: int, num_cols: int):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = _font(bold=True, size=11, color="white")
            cell.fill = _fill("dark")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

    def _style_data_row(self, ws, row: int, num_cols: int, zebra=False):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = _font()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
            if zebra:
                cell.fill = _fill("light_bg")

    def _set_col_widths(self, ws, widths: list[int]):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _add_title(self, ws, text: str, subtitle: str = ""):
        ws.merge_cells(f"A1:{get_column_letter(10)}1")
        ws.cell(row=1, column=1, value=text).font = _font(bold=True, size=16)
        ws.row_dimensions[1].height = 30
        if subtitle:
            ws.cell(row=2, column=1, value=subtitle).font = _font(color="gray")

    def _add_section(self, ws, row: int, title: str) -> int:
        ws.cell(row=row, column=1, value=title).font = _font(bold=True, size=12, color="primary")
        return row + 1

    def _write_table(self, ws, start_row: int, headers: list[str],
                     data: list[list], col_widths: Optional[list[int]] = None) -> int:
        """Write headers + data rows. Returns next available row."""
        for c, h in enumerate(headers, 1):
            ws.cell(row=start_row, column=c, value=h)
        self._style_header(ws, start_row, len(headers))

        for i, row_data in enumerate(data):
            r = start_row + 1 + i
            for c, val in enumerate(row_data, 1):
                ws.cell(row=r, column=c, value=val)
            self._style_data_row(ws, r, len(headers), zebra=(i % 2 == 1))

        if col_widths:
            self._set_col_widths(ws, col_widths)

        return start_row + 1 + len(data) + 1

    def _write_insights(self, ws, start_row: int, insights: list[str]) -> int:
        row = self._add_section(ws, start_row, "Key Insights")
        for text in insights:
            ws.cell(row=row, column=1, value=text).font = _font(color="gray")
            row += 1
        return row + 1

    def _add_bar_chart(self, ws, title: str, data_ref, cats_ref, anchor: str,
                       width=18, height=10):
        chart = BarChart()
        chart.type = "col"
        chart.title = title
        chart.style = 10
        chart.y_axis.title = None
        chart.x_axis.title = None
        chart.legend = None
        chart.width = width
        chart.height = height
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        # Apply brand color
        if chart.series:
            chart.series[0].graphicalProperties.solidFill = COLORS["primary"]
        ws.add_chart(chart, anchor)

    def _add_line_chart(self, ws, title: str, data_ref, cats_ref, anchor: str,
                        width=18, height=10):
        chart = LineChart()
        chart.title = title
        chart.style = 10
        chart.width = width
        chart.height = height
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        if chart.series:
            chart.series[0].graphicalProperties.line.solidFill = COLORS["primary"]
        ws.add_chart(chart, anchor)

    # ─── Sheet Builders ───

    def add_campaign_sheet(self, headers: list[str], data: list[list],
                           insights: list[str], col_widths: Optional[list[int]] = None):
        ws = self.wb.create_sheet("Campaign Overview")
        ws.sheet_properties.tabColor = COLORS["primary"]
        self._add_title(ws, "CAMPAIGN PERFORMANCE OVERVIEW",
                        f"Client: {self.client_name} | Period: {self.period}")
        next_row = self._write_table(ws, 4, headers, data, col_widths)
        # Charts will be built from actual data references
        if len(data) > 0:
            # Build chart from data in sheet
            chart_anchor_row = next_row + 1
            # (Agent adds charts after seeing the data)
        insight_row = next_row + 16  # leave room for charts
        self._write_insights(ws, insight_row, insights)
        self._insights["Campaign Overview"] = insights

    def add_creative_sheet(self, headers: list[str], data: list[list],
                           insights: list[str], col_widths: Optional[list[int]] = None):
        ws = self.wb.create_sheet("Creative Analysis")
        ws.sheet_properties.tabColor = COLORS["accent"]
        self._add_title(ws, "CREATIVE PERFORMANCE ANALYSIS",
                        f"Breakdown by ad creative | Period: {self.period}")
        next_row = self._write_table(ws, 4, headers, data, col_widths)
        insight_row = next_row + 16
        self._write_insights(ws, insight_row, insights)
        self._insights["Creative Analysis"] = insights

    def add_demographics_sheet(self, age_headers, age_data, gender_headers, gender_data,
                                insights: list[str], col_widths: Optional[list[int]] = None):
        ws = self.wb.create_sheet("Demographics")
        ws.sheet_properties.tabColor = COLORS["green"]
        self._add_title(ws, "DEMOGRAPHIC BREAKDOWN",
                        f"Age & Gender analysis | Period: {self.period}")
        row = self._add_section(ws, 4, "By Age Group")
        next_row = self._write_table(ws, row, age_headers, age_data, col_widths)
        next_row += 1
        row = self._add_section(ws, next_row, "By Gender")
        next_row = self._write_table(ws, row, gender_headers, gender_data, col_widths)
        insight_row = next_row + 16
        self._write_insights(ws, insight_row, insights)
        self._insights["Demographics"] = insights

    def add_custom_sheet(self, sheet_name: str, tab_color: str, title: str,
                         headers: list[str], data: list[list],
                         insights: list[str], col_widths: Optional[list[int]] = None):
        """Generic sheet for any breakdown (placement, time, device, geo, etc.)"""
        ws = self.wb.create_sheet(sheet_name)
        ws.sheet_properties.tabColor = COLORS.get(tab_color, tab_color)
        self._add_title(ws, title, f"Period: {self.period}")
        next_row = self._write_table(ws, 4, headers, data, col_widths)
        insight_row = next_row + 16
        self._write_insights(ws, insight_row, insights)
        self._insights[sheet_name] = insights

    def add_event_score_sheet(self, headers: list[str], data: list[list],
                               insights: list[str]):
        ws = self.wb.create_sheet("Event Score")
        ws.sheet_properties.tabColor = COLORS["purple"]
        self._add_title(ws, "EVENT SCORE ANALYSIS",
                        f"Pixel/CAPI event quality | Period: {self.period}")
        next_row = self._write_table(ws, 4, headers, data,
                                      [18, 10, 16, 14, 12, 40])
        ref_row = self._add_section(ws, next_row + 1, "Snapchat References")
        refs = [
            "Event Quality Score: https://businesshelp.snapchat.com/s/article/event-quality-score",
            "CAPI Setup: https://businesshelp.snapchat.com/s/article/conversions-api",
            "Snap Pixel: https://businesshelp.snapchat.com/s/article/snap-pixel",
        ]
        for text in refs:
            ws.cell(row=ref_row, column=1, value=text).font = _font(color="gray")
            ref_row += 1
        self._write_insights(ws, ref_row + 1, insights)
        self._insights["Event Score"] = insights

    def add_adset_changes_sheet(self, headers: list[str], data: list[list],
                                 insights: list[str]):
        ws = self.wb.create_sheet("Ad Set Changes")
        ws.sheet_properties.tabColor = COLORS["red"]
        self._add_title(ws, "AD SET CHANGE LOG ANALYSIS",
                        f"Change frequency & impact | Period: {self.period}")
        next_row = self._write_table(ws, 4, headers, data,
                                      [12, 22, 14, 16, 16, 16, 28])
        self._write_insights(ws, next_row + 1, insights)
        self._insights["Ad Set Changes"] = insights

    def add_audit_summary(self, executive_summary: str,
                          strengths: list[str], weaknesses: list[str],
                          root_causes: list[str], recommendations: list[str],
                          next_steps: list[str], references: Optional[list[str]] = None):
        """Build the final AUDIT SUMMARY sheet."""
        ws = self.wb.create_sheet("AUDIT SUMMARY")
        ws.sheet_properties.tabColor = COLORS["primary"]

        # Branded header
        ws.merge_cells("A1:F1")
        c = ws.cell(row=1, column=1, value=f"SNAPCHAT AD AUDIT — {self.client_name}")
        c.font = _font(bold=True, size=18, color="white")
        for col in range(1, 7):
            ws.cell(row=1, column=col).fill = _fill("primary")
        ws.row_dimensions[1].height = 40
        ws.cell(row=2, column=1,
                value=f"Prepared by: Allformance | Period: {self.period} | Date: {datetime.now().strftime('%d %b %Y')}").font = _font(color="gray")
        self._set_col_widths(ws, [50, 30, 20, 20, 20, 20])

        row = 4
        row = self._add_section(ws, row, "1. Executive Summary")
        ws.cell(row=row, column=1, value=executive_summary).font = _font(color="gray")
        row += 2

        row = self._add_section(ws, row, "2. Strengths")
        for s in strengths:
            ws.cell(row=row, column=1, value=f"✓ {s}").font = _font(color="green")
            row += 1
        row += 1

        row = self._add_section(ws, row, "3. Weaknesses & Issues")
        for w in weaknesses:
            ws.cell(row=row, column=1, value=f"✗ {w}").font = _font(color="red")
            row += 1
        row += 1

        row = self._add_section(ws, row, "4. Root Cause Analysis")
        for rc in root_causes:
            ws.cell(row=row, column=1, value=rc).font = _font(color="gray")
            row += 1
        row += 1

        row = self._add_section(ws, row, "5. Recommendations")
        for rec in recommendations:
            ws.cell(row=row, column=1, value=rec).font = _font(color="gray")
            row += 1
        row += 1

        if references:
            row = self._add_section(ws, row, "6. Snapchat References")
            for ref in references:
                ws.cell(row=row, column=1, value=ref).font = _font(color="gray")
                row += 1
            row += 1

        row = self._add_section(ws, row, f"{'7' if references else '6'}. Next Steps")
        for ns in next_steps:
            ws.cell(row=row, column=1, value=f"☐ {ns}").font = _font(color="gray")
            row += 1

    def save(self, filename: Optional[str] = None) -> str:
        if not filename:
            safe_name = self.client_name.replace(" ", "_").lower()
            date_str = datetime.now().strftime("%Y%m%d")
            filename = f"audit_{safe_name}_{date_str}.xlsx"
        path = os.path.join(self.output_dir, filename)
        self.wb.save(path)
        return path


# ─── Quick test ───
if __name__ == "__main__":
    r = AuditReport("Test Client", "Mar 1–26, 2026", "/Users/tolkozin/Audit")
    r.add_campaign_sheet(
        headers=["Campaign", "Spend", "CPA"],
        data=[["Test Campaign", "$1,000", "$5.00"]],
        insights=["Test insight"],
    )
    r.add_audit_summary(
        executive_summary="Test summary.",
        strengths=["Good CTR"],
        weaknesses=["High CPA"],
        root_causes=["Audience too broad"],
        recommendations=["Narrow targeting"],
        next_steps=["Review in 7 days"],
    )
    path = r.save("test_report.xlsx")
    print(f"Test report saved: {path}")
